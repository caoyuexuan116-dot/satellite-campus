from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


# =============================================================================
# 0. Global configuration
# =============================================================================
# This script is the data-cleaning code behind the current reported results.
# Pipeline:
#   1) read each school-level WoS Excel workbook;
#   2) split `Author Full Names` into author-level rows;
#   3) use WoS `Addresses` to keep only authors affiliated with the current school;
#   4) deduplicate same school-person-paper records by UT + person_id;
#   5) aggregate author-paper rows to a balanced person-year panel for Stata.
#
# Important: this is still a pilot sample and still ignores true name
# disambiguation when ORCID is unavailable. The fallback person_id is
# school + normalized full name, so homonyms inside a school can still merge.
WOS_DIR = Path(r"D:\数据\WOS数据（数据不超过10w条的学校）")
OUTPUT_DIR = Path("outputs")
YEAR_MIN = 2000
YEAR_MAX = 2022


# =============================================================================
# 1. Sample definition and affiliation aliases
# =============================================================================
# Fixed 5:5 pilot sample. Treatment schools opened same-city satellite campuses
# within the study window; strict controls have no local relocation-workbook
# campus record before or during the window.
#
# `address_aliases` are the English organization strings expected inside WoS
# `Addresses`. These aliases are deliberately conservative: if an author's
# address block does not contain one of these aliases, that author is dropped.
# When expanding to the full sample, each new school needs its own checked alias
# list, because WoS abbreviations vary by institution.
SCHOOLS = [
    {
        "school_cn": "北京理工大学",
        "wos_file": "Beijing Institute of Technology.xlsx",
        "treated": 1,
        "relo_year": 2007,
        "sample_role": "original_treat",
        "note": "良乡校区；同城新建多校区",
        "address_aliases": ["Beijing Inst Technol"],
    },
    {
        "school_cn": "中央财经大学",
        "wos_file": "Central University of Finance & Economics.xlsx",
        "treated": 1,
        "relo_year": 2009,
        "sample_role": "original_treat",
        "note": "沙河校区；同城新建多校区",
        "address_aliases": ["Cent Univ Finance & Econ", "Central Univ Finance & Econ"],
    },
    {
        "school_cn": "暨南大学",
        "wos_file": "Jinan University.xlsx",
        "treated": 1,
        "relo_year": 2014,
        "sample_role": "original_treat",
        "note": "番禺校区；同城新建多校区；不可误用 University of Jinan",
        "address_aliases": ["Jinan Univ"],
    },
    {
        "school_cn": "华东理工大学",
        "wos_file": "East China University of Science & Technology.xlsx",
        "treated": 1,
        "relo_year": 2007,
        "sample_role": "original_treat",
        "note": "奉贤校区；同城新建多校区",
        "address_aliases": ["East China Univ Sci & Technol"],
    },
    {
        "school_cn": "东华大学",
        "wos_file": "Donghua University.xlsx",
        "treated": 1,
        "relo_year": 2003,
        "sample_role": "replacement_treat",
        "note": "替代南京大学；松江校区；同城新建多校区",
        "address_aliases": ["Donghua Univ"],
    },
    {
        "school_cn": "中国传媒大学",
        "wos_file": "Communication University of China.xlsx",
        "treated": 0,
        "relo_year": None,
        "sample_role": "strict_control",
        "note": "本地搬迁表无记录；官网招生简介列北京定福庄东街一号",
        "address_aliases": ["Commun Univ China"],
    },
    {
        "school_cn": "东北林业大学",
        "wos_file": "Northeast Forestry University - China.xlsx",
        "treated": 0,
        "relo_year": None,
        "sample_role": "strict_control",
        "note": "本地搬迁表无记录；教育部章程核准书列哈尔滨市香坊区和兴路26号",
        "address_aliases": ["Northeast Forestry Univ"],
    },
    {
        "school_cn": "东北农业大学",
        "wos_file": "Northeast Agricultural University - China.xlsx",
        "treated": 0,
        "relo_year": None,
        "sample_role": "strict_control",
        "note": "本地搬迁表无记录；公开资料列哈尔滨市香坊区长江路600号",
        "address_aliases": ["Northeast Agr Univ", "Northeast Agric Univ"],
    },
    {
        "school_cn": "广西大学",
        "wos_file": "Guangxi University.xlsx",
        "treated": 0,
        "relo_year": None,
        "sample_role": "strict_control",
        "note": "本地搬迁表无记录；学校章程列南宁市大学东路100号",
        "address_aliases": ["Guangxi Univ"],
    },
    {
        "school_cn": "中央音乐学院",
        "wos_file": "Central Conservatory of Music.xlsx",
        "treated": 0,
        "relo_year": None,
        "sample_role": "strict_control",
        "note": "本地搬迁表无记录；官网联系方式列北京市西城区鲍家街43号；WOS 量较小",
        "address_aliases": ["Cent Conservatory Mus", "Central Conservatory Mus"],
    },
]

# Schools excluded from the pilot and the reason. `must_be_absent=1` is only for
# schools whose WOS file should not exist in the current directory; if such a file
# appears later, the script stops so the sample can be reviewed instead of
# silently continuing with stale assumptions.
EXCLUDED = [
    {
        "school_cn": "北京交通大学",
        "expected_wos_file": "Beijing Jiaotong University.xlsx",
        "reason": "用户收紧口径；对照组在窗口期和之前均不得有多校区，北交威海校区不再保留",
        "must_be_absent": 0,
    },
    {
        "school_cn": "对外经济贸易大学",
        "expected_wos_file": "University of International Business & Economics.xlsx",
        "reason": "用户收紧口径；对照组在窗口期和之前均不得有多校区，贸大青岛校区建设信息不再保留",
        "must_be_absent": 0,
    },
    {
        "school_cn": "北京工业大学",
        "expected_wos_file": "Beijing University of Technology.xlsx",
        "reason": "用户收紧口径；对照组在窗口期和之前均不得有多校区，北工大通州校区记录不再保留",
        "must_be_absent": 0,
    },
    {
        "school_cn": "中国政法大学",
        "expected_wos_file": "China University of Political Science & Law.xlsx",
        "reason": "用户收紧口径；对照组在窗口期和之前均不得有多校区，中国政法大学昌平校区记录不再保留",
        "must_be_absent": 0,
    },
    {
        "school_cn": "上海财经大学",
        "expected_wos_file": "Shanghai University of Finance & Economics.xlsx",
        "reason": "用户要求剔除；对照组在窗口期和之前均不得有多校区",
        "must_be_absent": 0,
    },
    {
        "school_cn": "南京大学",
        "expected_wos_file": "Nanjing University.xlsx",
        "reason": "当前 WOS 文件夹未找到对应文件，不能虚填",
        "must_be_absent": 1,
    },
    {
        "school_cn": "上海大学",
        "expected_wos_file": "Shanghai University.xlsx",
        "reason": "当前 WOS 文件夹未找到对应文件，不能虚填",
        "must_be_absent": 1,
    },
]


# =============================================================================
# 2. Output schemas
# =============================================================================
# Author-paper output: one row per WOS paper-author pair after splitting
# `Author Full Names` by semicolon.
#
# Caution: citation values are assigned to every retained author on a paper.
# This is an author-level crediting rule, not fractional counting.
PAPER_FIELDS = [
    "school_id",
    "school_cn",
    "school_code",
    "wos_english_name",
    "treated",
    "relo_year",
    "post",
    "year",
    "person_id",
    "person_id_type",
    "author_full_name",
    "matched_orcid",
    "ut",
    "doi",
    "article_title",
    "source_title",
    "document_type",
    "publication_type",
    "cites_wos_core",
    "cites_all_db",
]


# Person-year output: a balanced panel over 2000-2022 for every observed person.
# Years without observed papers are written as zero-output rows.
#
# Stata consumes this file directly. `did = treated * post`, where `post` turns
# on in each treated school's own relocation year.
PANEL_FIELDS = [
    "school_id",
    "school_cn",
    "treated",
    "relo_year",
    "year",
    "post",
    "did",
    "person_id",
    "person_id_type",
    "author_full_name",
    "pub_count",
    "cites_wos_core_sum",
    "cites_all_db_sum",
]


# =============================================================================
# 3. Generic parsing helpers
# =============================================================================
def clean_text(value: object) -> str:
    """Convert Excel values to stripped one-line strings."""
    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def parse_int(value: object) -> int | None:
    """Parse years and other integer-like Excel cells; invalid values become None."""
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except ValueError:
        return None


def parse_float(value: object) -> float:
    """Parse numeric outcomes; missing/invalid citation counts are treated as zero."""
    if value is None or value == "":
        return 0.0
    try:
        return float(str(value).strip())
    except ValueError:
        return 0.0


def split_semicolon(value: object) -> list[str]:
    """Split WoS semicolon-separated cells while dropping blank fragments."""
    text = clean_text(value)
    if not text:
        return []
    return [part.strip() for part in text.split(";") if part.strip()]


def norm_name(name: str) -> str:
    """Normalize names for conservative ORCID matching and fallback IDs."""
    return re.sub(r"[^a-z0-9]", "", name.lower())


def norm_org(text: str) -> str:
    """Normalize organization/address text for school alias matching."""
    return re.sub(r"[^a-z0-9]", "", text.lower())


def parse_orcids(value: object) -> dict[str, str]:
    """Parse WoS ORCID cells of the form 'Name/0000-....; Name/0000-....'."""
    pairs: dict[str, str] = {}
    for part in split_semicolon(value):
        if "/" not in part:
            continue
        name, orcid = part.rsplit("/", 1)
        orcid = orcid.strip()
        if re.fullmatch(r"\d{4}-\d{4}-\d{4}-[\dX]{4}", orcid):
            pairs[norm_name(name)] = orcid
    return pairs


def parse_current_school_authors(addresses: object, aliases: list[str], authors: list[str]) -> set[str]:
    """Return normalized author names whose WoS address block matches this school.

    WoS `Addresses` usually looks like:
    `[Author A; Author B] Univ Alias, Dept, City; [Author C] Other Univ, ...`.
    This parser keeps Author A/B only if their block text contains a current-school
    alias. If author-level brackets are absent, it only keeps a single-author paper
    when the address contains a school alias; multi-author unbracketed rows are
    dropped to avoid crediting outside collaborators to the current school.
    """
    text = clean_text(addresses)
    if not text:
        # No address information means we cannot verify current-school affiliation.
        # Drop all authors from the paper rather than over-crediting collaborators.
        return set()

    normalized_aliases = [norm_org(alias) for alias in aliases if alias]
    if not normalized_aliases:
        return set()

    current_school_authors: set[str] = set()
    # Extract author-address blocks. Example:
    #   [Li, A; Wang, B] Guangxi Univ, Coll Chem, Nanning; [Zhang, C] Other Univ...
    # Each block gets matched independently, so only authors in school-matching
    # blocks are retained.
    blocks = re.findall(r"\[([^\]]+)\]\s*([^\[]+?)(?=;\s*\[|$)", text)
    for block_authors, block_address in blocks:
        if not any(alias in norm_org(block_address) for alias in normalized_aliases):
            continue
        for block_author in split_semicolon(block_authors):
            current_school_authors.add(norm_name(block_author))

    if blocks:
        return current_school_authors

    # Rare fallback: some WoS rows have no bracketed author-address mapping.
    # For multi-author papers this is unsafe, because we cannot know which author
    # belongs to the school. We therefore only keep single-author rows if the raw
    # address text contains the school alias.
    if len(authors) == 1 and any(alias in norm_org(text) for alias in normalized_aliases):
        return {norm_name(authors[0])}

    return set()


def require_columns(headers: list[str], required: list[str], file_name: str) -> dict[str, int]:
    """Return a header index and stop early if any needed WoS field is missing."""
    index = {header: i for i, header in enumerate(headers)}
    missing = [col for col in required if col not in index]
    if missing:
        raise ValueError(f"{file_name} missing required columns: {missing}")
    return index


# =============================================================================
# 4. Pre-run validation and audit metadata
# =============================================================================
def validate_sample() -> None:
    """Fail early if the fixed sample is unbalanced or references wrong files."""
    missing_files = []
    for school in SCHOOLS:
        path = WOS_DIR / school["wos_file"]
        if not path.exists():
            missing_files.append(str(path))
    if missing_files:
        raise FileNotFoundError("Missing WOS files:\n" + "\n".join(missing_files))

    for excluded in EXCLUDED:
        path = WOS_DIR / excluded["expected_wos_file"]
        if int(excluded["must_be_absent"]) == 1 and path.exists():
            raise RuntimeError(f"Excluded file unexpectedly exists; review sample: {path}")

    treated_count = sum(int(s["treated"]) for s in SCHOOLS)
    control_count = len(SCHOOLS) - treated_count
    if treated_count != 5 or control_count != 5:
        raise RuntimeError(f"Expected 5:5 sample, got treated={treated_count}, control={control_count}")

    if (WOS_DIR / "University of Jinan.xlsx").exists() and not (WOS_DIR / "Jinan University.xlsx").exists():
        raise RuntimeError("Jinan University file is missing; do not use University of Jinan.")


def write_metadata() -> None:
    """Write small audit tables so sample choices can be inspected without code."""
    OUTPUT_DIR.mkdir(exist_ok=True)
    with (OUTPUT_DIR / "sample_school_mapping.csv").open("w", newline="", encoding="utf-8-sig") as f:
        fieldnames = [
            "school_id",
            "school_cn",
            "wos_file",
            "treated",
            "relo_year",
            "sample_role",
            "address_aliases",
            "note",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for i, school in enumerate(SCHOOLS, start=1):
            row = {**school, "school_id": i, "address_aliases": "; ".join(school["address_aliases"])}
            writer.writerow(row)

    with (OUTPUT_DIR / "excluded_schools.csv").open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["school_cn", "expected_wos_file", "reason", "must_be_absent"])
        writer.writeheader()
        writer.writerows(EXCLUDED)


# =============================================================================
# 5. Main cleaning pipeline
# =============================================================================
def main() -> None:
    # Validate fixed sample and write small tables that document sample choices.
    validate_sample()
    write_metadata()

    paper_path = OUTPUT_DIR / "wos_author_paper_sample.csv"
    panel_path = OUTPUT_DIR / "wos_author_year_panel_sample.csv"
    validation_path = OUTPUT_DIR / "validation_summary.txt"

    aggregates: dict[tuple[int, str, int], dict[str, float]] = defaultdict(
        lambda: {"pub_count": 0.0, "cites_wos_core_sum": 0.0, "cites_all_db_sum": 0.0}
    )
    # `aggregates` accumulates paper counts and citation sums by school-person-year.
    # `persons` stores one canonical row per observed school-person for panel expansion.
    persons: dict[tuple[int, str], dict[str, object]] = {}
    school_stats: list[dict[str, object]] = []

    # -------------------------------------------------------------------------
    # 5A. Build author-paper data and in-memory person-year aggregates
    # -------------------------------------------------------------------------
    # We stream each workbook in read-only mode. This matters because the WOS
    # school files can be large, and the generated author-paper CSV is also large.
    with paper_path.open("w", newline="", encoding="utf-8-sig") as f_paper:
        paper_writer = csv.DictWriter(f_paper, fieldnames=PAPER_FIELDS)
        paper_writer.writeheader()

        for school_id, school in enumerate(SCHOOLS, start=1):
            file_path = WOS_DIR / school["wos_file"]
            print(f"Reading {school_id}/{len(SCHOOLS)}: {school['school_cn']} ({school['wos_file']})", flush=True)
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            headers = [clean_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            idx = require_columns(
                headers,
                [
                    "code",
                    "EnglishName",
                    "Publication Type",
                    "Authors",
                    "Author Full Names",
                    "Article Title",
                    "Source Title",
                    "Document Type",
                    "Addresses",
                    "Affiliations",
                    "Publication Year",
                    "DOI",
                    "Times Cited, WoS Core",
                    "Times Cited, All Databases",
                    "ORCIDs",
                    "UT (Unique WOS ID)",
                ],
                school["wos_file"],
            )

            source_rows = 0
            in_window_rows = 0
            author_rows = 0
            current_school_author_rows = 0
            dropped_non_school_author_rows = 0
            duplicate_author_paper_rows = 0
            orcid_matches = 0
            seen_author_paper: set[tuple[int, str, str]] = set()

            for row in ws.iter_rows(min_row=2, values_only=True):
                source_rows += 1

                # Keep only papers inside the analysis window.
                year = parse_int(row[idx["Publication Year"]])
                if year is None or year < YEAR_MIN or year > YEAR_MAX:
                    continue

                # WoS stores full author names as a semicolon-separated list.
                # This first-stage test intentionally ignores full name disambiguation.
                authors = split_semicolon(row[idx["Author Full Names"]])
                if not authors:
                    continue

                in_window_rows += 1
                # Core correction relative to the earlier version: identify which
                # authors actually belong to the current school before writing rows.
                current_school_authors = parse_current_school_authors(
                    row[idx["Addresses"]],
                    list(school["address_aliases"]),
                    authors,
                )
                orcid_by_name = parse_orcids(row[idx["ORCIDs"]])
                cites_wos_core = parse_float(row[idx["Times Cited, WoS Core"]])
                cites_all_db = parse_float(row[idx["Times Cited, All Databases"]])
                ut = clean_text(row[idx["UT (Unique WOS ID)"]])
                relo_year = school["relo_year"]
                treated = int(school["treated"])
                post = int(treated == 1 and relo_year is not None and year >= int(relo_year))

                for author in authors:
                    normalized = norm_name(author)
                    if normalized not in current_school_authors:
                        # This author is a collaborator or cannot be verified as a
                        # current-school author from Addresses, so do not count them.
                        dropped_non_school_author_rows += 1
                        continue

                    matched_orcid = orcid_by_name.get(normalized, "")
                    # Use ORCID only when the ORCID name exactly matches this author
                    # after normalization; otherwise fall back to school + full name.
                    if matched_orcid:
                        person_id = f"orcid:{matched_orcid}"
                        person_id_type = "orcid"
                        orcid_matches += 1
                    else:
                        person_id = f"school_name:{school_id}:{normalized}"
                        person_id_type = "school_name"

                    author_paper_key = (school_id, person_id, ut)
                    if author_paper_key in seen_author_paper:
                        # Prevent duplicate school-person-paper rows from repeated
                        # address blocks or duplicate WOS rows.
                        duplicate_author_paper_rows += 1
                        continue
                    seen_author_paper.add(author_paper_key)

                    current_school_author_rows += 1
                    author_rows += 1
                    key = (school_id, person_id, year)
                    # Counts and citation sums are author-level credit: each listed
                    # author receives one publication count and the paper citation value.
                    aggregates[key]["pub_count"] += 1
                    aggregates[key]["cites_wos_core_sum"] += cites_wos_core
                    aggregates[key]["cites_all_db_sum"] += cites_all_db
                    persons[(school_id, person_id)] = {
                        "school_id": school_id,
                        "school_cn": school["school_cn"],
                        "treated": treated,
                        "relo_year": "" if relo_year is None else relo_year,
                        "person_id": person_id,
                        "person_id_type": person_id_type,
                        "author_full_name": author,
                    }

                    # Author-paper record used for audits and possible downstream
                    # checks. The regression uses the aggregated person-year file.
                    paper_writer.writerow(
                        {
                            "school_id": school_id,
                            "school_cn": school["school_cn"],
                            "school_code": clean_text(row[idx["code"]]),
                            "wos_english_name": clean_text(row[idx["EnglishName"]]),
                            "treated": treated,
                            "relo_year": "" if relo_year is None else relo_year,
                            "post": post,
                            "year": year,
                            "person_id": person_id,
                            "person_id_type": person_id_type,
                            "author_full_name": author,
                            "matched_orcid": matched_orcid,
                            "ut": ut,
                            "doi": clean_text(row[idx["DOI"]]),
                            "article_title": clean_text(row[idx["Article Title"]]),
                            "source_title": clean_text(row[idx["Source Title"]]),
                            "document_type": clean_text(row[idx["Document Type"]]),
                            "publication_type": clean_text(row[idx["Publication Type"]]),
                            "cites_wos_core": cites_wos_core,
                            "cites_all_db": cites_all_db,
                        }
                    )

            school_stats.append(
                {
                    "school_cn": school["school_cn"],
                    "wos_file": school["wos_file"],
                    "source_rows": source_rows,
                    "in_window_paper_rows": in_window_rows,
                    "author_paper_rows": author_rows,
                    "current_school_author_rows": current_school_author_rows,
                    "dropped_non_school_author_rows": dropped_non_school_author_rows,
                    "duplicate_author_paper_rows": duplicate_author_paper_rows,
                    "orcid_matches": orcid_matches,
                }
            )
            wb.close()
            print(
                f"Finished {school['school_cn']}: source_rows={source_rows}, "
                f"in_window_paper_rows={in_window_rows}, current_school_author_rows={current_school_author_rows}, "
                f"dropped_non_school_author_rows={dropped_non_school_author_rows}",
                flush=True,
            )

    # -------------------------------------------------------------------------
    # 5B. Expand to balanced person-year panel
    # -------------------------------------------------------------------------
    # Every observed person receives one row for every year in 2000-2022.
    # This makes zero-output years explicit and lets Stata run person fixed effects
    # on a strongly balanced panel.
    print(f"Writing balanced person-year panel for {len(persons)} persons", flush=True)
    with panel_path.open("w", newline="", encoding="utf-8-sig") as f_panel:
        panel_writer = csv.DictWriter(f_panel, fieldnames=PANEL_FIELDS)
        panel_writer.writeheader()

        for person_key in sorted(persons):
            person = persons[person_key]
            treated = int(person["treated"])
            relo_year_value = person["relo_year"]
            relo_year = int(relo_year_value) if relo_year_value != "" else None
            # Expand every observed person to all calendar years in the window.
            # This makes absence of publications explicit instead of dropping those years.
            for year in range(YEAR_MIN, YEAR_MAX + 1):
                post = int(treated == 1 and relo_year is not None and year >= relo_year)
                did = treated * post
                vals = aggregates.get((int(person["school_id"]), str(person["person_id"]), year))
                panel_writer.writerow(
                    {
                        **person,
                        "year": year,
                        "post": post,
                        "did": did,
                        "pub_count": int(vals["pub_count"]) if vals else 0,
                        "cites_wos_core_sum": vals["cites_wos_core_sum"] if vals else 0,
                        "cites_all_db_sum": vals["cites_all_db_sum"] if vals else 0,
                    }
                )

    # -------------------------------------------------------------------------
    # 5C. Write validation summary
    # -------------------------------------------------------------------------
    # This summary is intentionally small enough to commit to GitHub. It records
    # how many author rows were retained, dropped as outside-school authors, and
    # removed as duplicate UT-person pairs.
    with validation_path.open("w", encoding="utf-8-sig") as f:
        f.write("Sample validation summary\n")
        f.write("=========================\n")
        f.write(f"Window: {YEAR_MIN}-{YEAR_MAX}\n")
        f.write("Sample balance: treated=5, control=5\n")
        f.write("Panel construction: balanced person-year rows for every observed person over 2000-2022.\n")
        f.write("Excluded schools: 南京大学 and 上海大学 because expected WOS files are absent; 北京交通大学、对外经济贸易大学、北京工业大学、上海财经大学、中国政法大学 because they fail the strict control rule.\n")
        f.write("Control definition: no multicampus before or during the study window.\n")
        f.write("Strict controls are selected from schools with no record in the local campus relocation workbook and with WOS data.\n\n")
        for stat in school_stats:
            f.write(
                "{school_cn}\t{wos_file}\tsource_rows={source_rows}\t"
                "in_window_paper_rows={in_window_paper_rows}\tauthor_paper_rows={author_paper_rows}\t"
                "current_school_author_rows={current_school_author_rows}\t"
                "dropped_non_school_author_rows={dropped_non_school_author_rows}\t"
                "duplicate_author_paper_rows={duplicate_author_paper_rows}\t"
                "orcid_matches={orcid_matches}\n".format(**stat)
            )
        f.write(f"\nUnique persons: {len(persons)}\n")
        f.write(f"Person-year rows: {len(persons) * (YEAR_MAX - YEAR_MIN + 1)}\n")
        f.write(f"Author-paper CSV: {paper_path}\n")
        f.write(f"Person-year CSV: {panel_path}\n")

    print(f"Wrote {paper_path}")
    print(f"Wrote {panel_path}")
    print(f"Wrote {validation_path}")


if __name__ == "__main__":
    main()
